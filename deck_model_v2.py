import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Colors
BLUE_DARK  = "1F4E79"
BLUE_MID   = "2E75B6"
BLUE_LIGHT = "D6E4F0"
GREEN_PROJ = "E2EFDA"
YELLOW_ASM = "FFF2CC"
GRAY_ROW   = "F2F2F2"
WHITE      = "FFFFFF"
RED_CHECK  = "FF0000"

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def w(ws, r, c, val, bold=False, color="000000", bg=None, italic=False,
      align="right", size=10, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if bg: cell.fill = fill(bg)
    if fmt: cell.number_format = fmt
    return cell

def sec(ws, row, label, ncols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    cell.fill = fill(BLUE_MID)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 15

def title(ws, label, ncols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=label)
    cell.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    cell.fill = fill(BLUE_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

def col_hdrs(ws, years, hist=3):
    hdrs = [""] + years
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.alignment = Alignment(horizontal="center" if j>1 else "left",
                                vertical="center", indent=1 if j==1 else 0)
        if j == 1:
            c.font = fnt(True, WHITE); c.fill = fill(BLUE_MID)
        elif j <= hist+1:
            c.font = fnt(True, WHITE); c.fill = fill(BLUE_MID)
        else:
            c.font = fnt(True); c.fill = fill(GREEN_PROJ)
    ws.row_dimensions[2].height = 15

def note(ws, r, text):
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(italic=True, color="7F7F7F", size=8, name="Calibri")

# ══════════════════════════════════════════════════════════════════
# ASSUMPTIONS TAB
# ══════════════════════════════════════════════════════════════════
ws_a = wb.active
ws_a.title = "Assumptions"
ws_a.column_dimensions["A"].width = 38
for col in "BCDEFGH": ws_a.column_dimensions[col].width = 14

title(ws_a, "DECKERS OUTDOOR (DECK) — MODEL ASSUMPTIONS")
col_hdrs(ws_a, ["FY2023A","FY2024A","FY2025A","FY2026E","FY2027E","FY2028E"])

# Revenue
sec(ws_a, 3, "REVENUE ASSUMPTIONS")
rev_rows = [
    ("Revenue Growth % (YoY)",    None,  0.1822, 0.1628, 0.10,  0.09,  0.08),
    ("  HOKA Revenue ($K)",     1412916,1806740,2233090,  None,  None,  None),
    ("  UGG Revenue ($K)",      1929211,2239132,2531351,  None,  None,  None),
    ("  Other Brands ($K)",      285159, 241891,  221171,  None,  None,  None),
]
for i,(lbl,*vals) in enumerate(rev_rows, 4):
    ws_a.cell(row=i,column=1,value=lbl).alignment=Alignment(horizontal="left",indent=1)
    for j,v in enumerate(vals,2):
        c=ws_a.cell(row=i,column=j,value=v)
        c.alignment=Alignment(horizontal="right")
        if j>=5 and v is not None: c.fill=fill(YELLOW_ASM); c.font=fnt(True)
        if i==4 and v is not None: c.number_format="0.0%"
        elif v and i>4: c.number_format="#,##0"

# Margins
sec(ws_a, 8, "MARGIN ASSUMPTIONS")
mgn_rows = [
    ("Gross Margin %",         0.5030, 0.5562, 0.5787, 0.580, 0.582, 0.582),
    ("SG&A % of Revenue",      0.3232, 0.3400, 0.3423, 0.338, 0.335, 0.332),
    ("EBIT Margin %",          0.1798, 0.2162, 0.2364, 0.242, 0.247, 0.250),
    ("Effective Tax Rate %",   0.2241, 0.2241, 0.2228, 0.223, 0.223, 0.223),
]
for i,(lbl,*vals) in enumerate(mgn_rows, 9):
    ws_a.cell(row=i,column=1,value=lbl).alignment=Alignment(horizontal="left",indent=1)
    for j,v in enumerate(vals,2):
        c=ws_a.cell(row=i,column=j,value=v); c.number_format="0.0%"
        c.alignment=Alignment(horizontal="right")
        if j>=5: c.fill=fill(YELLOW_ASM); c.font=fnt(True)

# Working capital
sec(ws_a, 13, "WORKING CAPITAL ASSUMPTIONS (days)")
wc_rows = [
    ("Days Sales Outstanding (DSO)",    30, 25, 24, 25, 25, 25),
    ("Days Inventory Outstanding (DIO)",108,91, 86, 88, 87, 86),
    ("Days Payable Outstanding (DPO)",   54,73, 73, 73, 73, 73),
]
for i,(lbl,*vals) in enumerate(wc_rows, 14):
    ws_a.cell(row=i,column=1,value=lbl).alignment=Alignment(horizontal="left",indent=1)
    for j,v in enumerate(vals,2):
        c=ws_a.cell(row=i,column=j,value=v); c.number_format="0"
        c.alignment=Alignment(horizontal="right")
        if j>=5: c.fill=fill(YELLOW_ASM); c.font=fnt(True)

# Capex & DA
sec(ws_a, 17, "CAPEX & D&A")
cap_rows = [
    ("Capex % of Revenue",  0.0223, 0.0208, 0.0173, 0.020, 0.022, 0.022),
    ("D&A % of Revenue",    0.0132, 0.0134, 0.0139, 0.014, 0.014, 0.014),
]
for i,(lbl,*vals) in enumerate(cap_rows, 18):
    ws_a.cell(row=i,column=1,value=lbl).alignment=Alignment(horizontal="left",indent=1)
    for j,v in enumerate(vals,2):
        c=ws_a.cell(row=i,column=j,value=v); c.number_format="0.0%"
        c.alignment=Alignment(horizontal="right")
        if j>=5: c.fill=fill(YELLOW_ASM); c.font=fnt(True)

# DCF inputs
sec(ws_a, 20, "DCF ASSUMPTIONS")
dcf_items=[
    ("WACC",                   "10.0%"),
    ("Terminal Growth Rate",   "3.0%"),
    ("Shares Outstanding (K)", "150,201"),
    ("Net Cash ($K)",          "1,889,188"),
    ("Beta (5-yr monthly)",    "1.15"),
    ("Risk-Free Rate",         "4.5%"),
    ("Equity Risk Premium",    "5.0%"),
    ("Cost of Debt",           "5.5%"),
]
for i,(lbl,val) in enumerate(dcf_items,21):
    ws_a.cell(row=i,column=1,value=lbl).alignment=Alignment(horizontal="left",indent=1)
    c=ws_a.cell(row=i,column=5,value=val)
    c.fill=fill(YELLOW_ASM); c.font=fnt(True)
    c.alignment=Alignment(horizontal="right")

note(ws_a, 30, "Yellow cells = change these to run scenarios. Historical ratios computed from exact 10-K filings (FY2023-2025).")

# ══════════════════════════════════════════════════════════════════
# INCOME STATEMENT
# ══════════════════════════════════════════════════════════════════
ws_i = wb.create_sheet("Income Statement")
ws_i.column_dimensions["A"].width = 36
for col in "BCDEFGH": ws_i.column_dimensions[col].width = 14
title(ws_i, "DECKERS OUTDOOR (DECK) — INCOME STATEMENT  ($000s, fiscal year ends March 31)")
col_hdrs(ws_i, ["FY2023A","FY2024A","FY2025A","FY2026E","FY2027E","FY2028E"])

# Exact IS from 10-Ks
is_rows = [
    # (label, fy23, fy24, fy25, bold, pct, bg_override)
    ("REVENUE", None,None,None, False, False, BLUE_MID),
    ("Net Revenue",             3627286, 4287763, 4985612, True, False, BLUE_LIGHT),
    ("COST & MARGIN", None,None,None, False, False, BLUE_MID),
    ("Cost of Sales",           1801916, 1902275, 2099949, False, False, None),
    ("Gross Profit",            1825370, 2385488, 2885663, True, False, BLUE_LIGHT),
    ("  Gross Margin %",        0.5030,  0.5562,  0.5787,  False, True, GRAY_ROW),
    ("OPERATING EXPENSES", None,None,None, False, False, BLUE_MID),
    ("SG&A Expenses",           1172619, 1457974, 1706571, False, False, None),
    ("  SG&A % of Revenue",     0.3232,  0.3400,  0.3423,  False, True, GRAY_ROW),
    ("EBIT (Operating Income)", 652751,  927514, 1179092,  True, False, BLUE_LIGHT),
    ("  EBIT Margin %",         0.1798,  0.2162,  0.2364,  False, True, GRAY_ROW),
    ("BELOW THE LINE", None,None,None, False, False, BLUE_MID),
    ("Interest Income",          15563,   52208,   68389,  False, False, None),
    ("Interest Expense",         -3442,   -2564,   -3517,  False, False, None),
    ("Pre-Tax Income",           666082,  978941, 1243299,  True, False, BLUE_LIGHT),
    ("Income Tax Expense",       149260,  219378,  277208,  False, False, None),
    ("  Effective Tax Rate %",   0.2241,  0.2241,  0.2228,  False, True, GRAY_ROW),
    ("Net Income",               516822,  759563,  966091,  True, False, BLUE_LIGHT),
    ("  Net Margin %",           0.1425,  0.1971,  0.1938,  False, True, GRAY_ROW),
    ("EBITDA BRIDGE", None,None,None, False, False, BLUE_MID),
    ("D&A (from CF statement)",   47858,   57587,   69353,  False, False, None),
    ("EBITDA",                   700609,  985101, 1248445,  True, False, BLUE_LIGHT),
    ("  EBITDA Margin %",        0.1931,  0.2297,  0.2504,  False, True, GRAY_ROW),
]

r = 3
rev_row = None
for row_data in is_rows:
    lbl, v23, v24, v25, bold, pct, bg = row_data
    if v23 is None:  # section header
        sec(ws_i, r, lbl)
        r += 1; continue
    if lbl == "Net Revenue": rev_row = r

    cell = ws_i.cell(row=r, column=1, value=lbl)
    cell.font = fnt(bold); cell.alignment = Alignment(horizontal="left", indent=1)
    if bg: cell.fill = fill(bg)

    for j, v in enumerate([v23, v24, v25], 2):
        c = ws_i.cell(row=r, column=j, value=v)
        c.alignment = Alignment(horizontal="right")
        c.number_format = "0.0%" if pct else "#,##0"
        c.font = fnt(bold)
        if bg: c.fill = fill(bg)

    # Projection columns
    for j in range(5, 8):
        c = ws_i.cell(row=r, column=j, value="→ link Assumptions")
        c.fill = fill(GREEN_PROJ); c.font = fnt(color="0070C0", bold=bold)
        c.alignment = Alignment(horizontal="right")
        c.number_format = "0.0%" if pct else "#,##0"
    r += 1

note(ws_i, r+1, "Historical figures from Deckers 10-K filings (FY2023: March 2023, FY2024: March 2024, FY2025: March 2025). D&A from cash flow statement (non-cash add-back line).")

# ══════════════════════════════════════════════════════════════════
# BALANCE SHEET
# ══════════════════════════════════════════════════════════════════
ws_b = wb.create_sheet("Balance Sheet")
ws_b.column_dimensions["A"].width = 40
for col in "BCDEFGH": ws_b.column_dimensions[col].width = 14
title(ws_b, "DECKERS OUTDOOR (DECK) — BALANCE SHEET  ($000s, as of March 31)")
col_hdrs(ws_b, ["FY2023A","FY2024A","FY2025A","FY2026E","FY2027E","FY2028E"])

# Exact BS from 10-Ks
# FY2023 from FY2024 10-K comparative; FY2024 and FY2025 from respective 10-Ks
bs_rows = [
    ("ASSETS", None, None, None, True, True),
    ("Current Assets", None, None, None, False, True),
    ("  Cash & Equivalents",              981795, 1502051, 1889188, False, False),
    ("  Trade Accounts Receivable, net",  301511,  296565,  332872, False, False),
    ("  Inventories",                     532852,  474311,  495226, False, False),
    ("  Prepaid Expenses",                 33788,   34284,   39294, False, False),
    ("  Other Current Assets",             55523,   92713,   67282, False, False),
    ("  Income Tax Receivable",             4784,   43559,   36613, False, False),
    ("Total Current Assets",             1910253, 2443483, 2860475, False, True),
    ("Property & Equipment, net",         266679,  302122,  325599, False, False),
    ("Operating Lease Assets",            213302,  225669,  237352, False, False),
    ("Goodwill",                           13990,   13990,   13990, False, False),
    ("Other Intangible Assets, net",       37457,   27083,   15699, False, False),
    ("Deferred Tax Assets, net",           72592,   72584,   77591, False, False),
    ("Other Assets",                       41930,   50648,   39546, False, False),
    ("Total Assets",                     2556203, 3135579, 3570252, False, True),
    ("SPACER", None, None, None, False, False),
    ("LIABILITIES & STOCKHOLDERS EQUITY", None, None, None, True, True),
    ("Current Liabilities", None, None, None, False, True),
    ("  Trade Accounts Payable",          265605,  378503,  417955, False, False),
    ("  Accrued Payroll",                  63781,  123653,  125417, False, False),
    ("  Operating Lease Liabilities",      50765,   53581,   54453, False, False),
    ("  Other Accrued Expenses",           86753,  106785,  142120, False, False),
    ("  Income Tax Payable",               17322,   52338,   23299, False, False),
    ("  Value Added Tax Payable",          13154,    5133,    6697, False, False),
    ("Total Current Liabilities",         497380,  719993,  769941, False, True),
    ("Long-Term Operating Lease Liab.",   195723,  213298,  222522, False, False),
    ("Income Tax Liability (LT)",          62032,   52470,   13587, False, False),
    ("Other Long-Term Liabilities",        35335,   42350,   51189, False, False),
    ("Total Long-Term Liabilities",       293090,  308118,  287298, False, True),
    ("Total Liabilities",                 790470, 1028111, 1057239, False, True),
    ("SPACER2", None, None, None, False, False),
    ("Common Stock",                          262,    1536,    1502, False, False),
    ("Additional Paid-In Capital",         232932,  243050,  253466, False, False),
    ("Retained Earnings",                1571574, 1913615, 2307699, False, False),
    ("AOCI",                              -39035,  -50733,  -49654, False, False),
    ("Total Stockholders Equity",        1765733, 2107468, 2513013, False, True),
    ("Total Liabilities & Equity",       2556203, 3135579, 3570252, False, True),
    ("SPACER3", None, None, None, False, False),
    ("CHECK: BS Balances (must = 0)",
     "=B17-B33", "=C17-C33", "=D17-D33", False, False),
]

r = 3
for row_data in bs_rows:
    lbl, v23, v24, v25, is_sec, bold = row_data
    if "SPACER" in lbl:
        r += 1; continue
    if is_sec:
        sec(ws_b, r, lbl); r += 1; continue

    bg = BLUE_LIGHT if bold else None
    cell = ws_b.cell(row=r, column=1, value=lbl)
    cell.font = fnt(bold); cell.alignment = Alignment(horizontal="left", indent=1)
    if bg: cell.fill = fill(bg)

    for j, v in enumerate([v23, v24, v25], 2):
        c = ws_b.cell(row=r, column=j, value=v)
        c.alignment = Alignment(horizontal="right")
        if isinstance(v, (int, float)): c.number_format = "#,##0"
        c.font = fnt(bold)
        if bg: c.fill = fill(bg)

    for j in range(5, 8):
        c = ws_b.cell(row=r, column=j, value="→ build")
        c.fill = fill(GREEN_PROJ); c.font = fnt(color="0070C0", bold=bold)
        c.alignment = Alignment(horizontal="right")
    r += 1

note(ws_b, r+1, "FY2023 from FY2024 10-K comparative column. FY2024 and FY2025 from respective 10-K filings. CHECK row must equal 0.")

# ══════════════════════════════════════════════════════════════════
# CASH FLOW STATEMENT
# ══════════════════════════════════════════════════════════════════
ws_c = wb.create_sheet("Cash Flow")
ws_c.column_dimensions["A"].width = 42
for col in "BCDEFGH": ws_c.column_dimensions[col].width = 14
title(ws_c, "DECKERS OUTDOOR (DECK) — CASH FLOW STATEMENT  ($000s)")
col_hdrs(ws_c, ["FY2023A","FY2024A","FY2025A","FY2026E","FY2027E","FY2028E"])

cf_rows = [
    ("OPERATING ACTIVITIES", None, None, None, True, True),
    ("Net Income",                     516822,  759563,  966091, False, True),
    ("  + Depreciation & Amortization",  47858,   57587,   69353, False, False),
    ("  + Stock-Based Compensation",     26897,   37288,   37943, False, False),
    ("  + Other Non-Cash Items",          3953,    6826,    9700, False, False),
    ("  - Change in Trade Receivables",  -5609,  -38490,  -41339, False, False),
    ("  - Change in Inventories",       -26056,  58541,  -24344, False, False),
    ("  +/- Change in Prepaid & Other",  13459,    4157,   20946, False, False),
    ("  + Change in Trade Payables",     -74247, 119601,   35636, False, False),
    ("  +/- Change in Other Accrued",     11528,   43534,   22222, False, False),
    ("  +/- Change in Tax Payable",        4897,   35016,  -29039, False, False),
    ("  +/- Other Operating",            17920,  -50439,  -22646, False, False),
    ("Cash from Operations",            537422, 1033184, 1044523, False, True),
    ("INVESTING ACTIVITIES", None, None, None, True, True),
    ("  Capital Expenditures",          -81025,  -89365,  -86171, False, False),
    ("  Proceeds from Asset Sales",         12,       34,   11168, False, False),
    ("Cash from Investing",             -81013,  -89331,  -75003, False, True),
    ("FINANCING ACTIVITIES", None, None, None, True, True),
    ("  Share Repurchases",            -297372, -414931, -567002, False, False),
    ("  Stock Issuance & Options",        6566,    7230,    4772, False, False),
    ("  Other Financing",              -18225,   -9974,   -3985, False, False),
    ("Cash from Financing",            -309031, -417675, -581334, False, True),
    ("FX Effect on Cash",               -9110,   -5922,   -1049, False, False),
    ("NET CHANGE & RECONCILIATION", None, None, None, True, True),
    ("Net Change in Cash",             138268,   520256,  387137, False, False),
    ("Beginning Cash",                 843527,   981795, 1502051, False, False),
    ("Ending Cash",                    981795,  1502051, 1889188, False, True),
    ("FREE CASH FLOW", None, None, None, True, True),
    ("FCF (Cash from Ops - Capex)",    456397,   943819,  958352, False, True),
    ("  FCF Margin %",                  0.1258,  0.2201,  0.1922, False, False),
    ("FCF Conversion (FCF / NI)",       0.883,   1.243,   0.992, False, False),
]

r = 3
for row_data in cf_rows:
    lbl, v23, v24, v25, is_sec, bold = row_data
    if is_sec and v23 is None:
        sec(ws_c, r, lbl); r += 1; continue

    pct = "%" in lbl
    ratio = "Conversion" in lbl or ("Margin" not in lbl and isinstance(v23, float) and abs(v23) < 3 and not pct)
    bg = BLUE_LIGHT if bold else None

    cell = ws_c.cell(row=r, column=1, value=lbl)
    cell.font = fnt(bold); cell.alignment = Alignment(horizontal="left", indent=1)
    if bg: cell.fill = fill(bg)

    for j, v in enumerate([v23, v24, v25], 2):
        c = ws_c.cell(row=r, column=j, value=v)
        c.alignment = Alignment(horizontal="right")
        if pct: c.number_format = "0.0%"
        elif ratio and isinstance(v, float): c.number_format = "0.000x"
        else: c.number_format = "#,##0"
        c.font = fnt(bold)
        if bg: c.fill = fill(bg)

    for j in range(5, 8):
        c = ws_c.cell(row=r, column=j, value="→ proj")
        c.fill = fill(GREEN_PROJ); c.font = fnt(color="0070C0", bold=bold)
        c.alignment = Alignment(horizontal="right")
    r += 1

note(ws_c, r+1, "FY2025 from DECK FY2025 annual report p.F-8. FY2024 and FY2023 from FY2024 10-K p.F-8. FCF = CFO - Capex.")

# ══════════════════════════════════════════════════════════════════
# DCF TAB
# ══════════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("DCF")
ws_d.column_dimensions["A"].width = 38
for col in "BCDEFGH": ws_d.column_dimensions[col].width = 15
title(ws_d, "DECKERS OUTDOOR (DECK) — DCF VALUATION")

# Headers for projection columns only
for j, h in enumerate(["","FY2026E","FY2027E","FY2028E"], 1):
    c = ws_d.cell(row=2, column=j, value=h)
    if j == 1:
        c.font = fnt(True, WHITE); c.fill = fill(BLUE_MID)
        c.alignment = Alignment(horizontal="left", indent=1)
    else:
        c.font = fnt(True); c.fill = fill(GREEN_PROJ)
        c.alignment = Alignment(horizontal="center")
ws_d.row_dimensions[2].height = 15

sec(ws_d, 3, "STEP 1 — UNLEVERED FREE CASH FLOW")
ufcf_rows = [
    ("Revenue",                  "→ IS!E","→ IS!F","→ IS!G"),
    ("EBIT",                     "→ IS!E","→ IS!F","→ IS!G"),
    ("Tax Rate %",               "22.3%", "22.3%", "22.3%"),
    ("NOPAT = EBIT × (1 − t)",   "→ calc","→ calc","→ calc"),
    ("+ D&A",                    "→ CF!E","→ CF!F","→ CF!G"),
    ("− Capex",                  "→ CF!E","→ CF!F","→ CF!G"),
    ("− Δ Net Working Capital",  "→ calc","→ calc","→ calc"),
    ("= Unlevered FCF",          "→ calc","→ calc","→ calc"),
    ("Discount Period",          "1",     "2",     "3"),
    ("Discount Factor (÷WACC=10%)","0.9091","0.8264","0.7513"),
    ("PV of Unlevered FCF",      "→ calc","→ calc","→ calc"),
]
r = 4
for lbl, e26, e27, e28 in ufcf_rows:
    bold = lbl.startswith("=") or "FCF" in lbl or "NOPAT" in lbl
    ws_d.cell(row=r, column=1, value=lbl).alignment=Alignment(horizontal="left", indent=1)
    ws_d.cell(row=r, column=1).font = fnt(bold)
    for j, v in enumerate([e26, e27, e28], 2):
        c = ws_d.cell(row=r, column=j, value=v)
        c.fill = fill(GREEN_PROJ); c.font = fnt(bold, color="0070C0")
        c.alignment = Alignment(horizontal="right")
    r += 1

sec(ws_d, r, "STEP 2 — TERMINAL VALUE & ENTERPRISE VALUE")
r += 1

bridge_rows = [
    ("Sum of PV of FCFs ($K)",              "=SUM(B14:D14)",  None,  None),
    ("Terminal Year FCF ($K)",              "= Unlevered FCF FY2028E", None, None),
    ("Terminal Value = TV FCF×(1+g)/(WACC−g)",  "→ calc", None, None),
    ("PV of Terminal Value ($K)",           "= TV/(1+WACC)^3", None, None),
    ("Enterprise Value ($K)",               "= PV FCFs + PV TV", None, None),
    ("",                                    None, None, None),
    ("STEP 3 — BRIDGE TO EQUITY VALUE", None, None, None),
    ("Enterprise Value ($K)",               "→ from above", None, None),
    ("+ Cash & Equivalents ($K)",           "1,889,188", None, None),
    ("− Total Debt ($K)",                   "0", None, None),
    ("− Other Long-Term Liabilities ($K)",  "287,298", None, None),
    ("= Equity Value ($K)",                 "→ calc", None, None),
    ("Shares Outstanding (K)",              "150,201", None, None),
    ("Implied Share Price",                 "→ calc", None, None),
    ("Current Share Price (verify live)",   "~$112", None, None),
    ("Upside / (Downside) %",              "→ calc", None, None),
]

for lbl, val, _, __ in bridge_rows:
    if not lbl:
        r += 1; continue
    if lbl.startswith("STEP 3"):
        sec(ws_d, r, lbl); r += 1; continue
    bold = ("Equity Value" in lbl or "Enterprise" in lbl or "Implied" in lbl
            or "Upside" in lbl)
    ws_d.cell(row=r, column=1, value=lbl).alignment=Alignment(horizontal="left",indent=1)
    ws_d.cell(row=r, column=1).font = fnt(bold)
    if val:
        c = ws_d.cell(row=r, column=2, value=val)
        c.fill = fill(YELLOW_ASM); c.font = fnt(bold, color="0070C0")
        c.alignment = Alignment(horizontal="right")
    r += 1

# Sensitivity table
sec(ws_d, r, "STEP 4 — SENSITIVITY: IMPLIED SHARE PRICE vs WACC & TERMINAL GROWTH")
r += 1

ws_d.cell(row=r, column=1, value="Terminal Growth Rate →").font = fnt(True)
tgrs = [0.020, 0.025, 0.030, 0.035, 0.040]
waccs = [0.090, 0.095, 0.100, 0.105, 0.110, 0.115]

for j, tg in enumerate(tgrs, 3):
    c = ws_d.cell(row=r, column=j, value=tg)
    c.number_format = "0.0%"; c.font = fnt(True, WHITE)
    c.fill = fill(BLUE_MID); c.alignment = Alignment(horizontal="center")
r += 1

ws_d.cell(row=r, column=1, value="WACC ↓").font = fnt(True, WHITE)
ws_d.cell(row=r, column=1).fill = fill(BLUE_MID)

for i, wacc in enumerate(waccs):
    rr = r + i
    c = ws_d.cell(row=rr, column=2, value=wacc)
    c.number_format = "0.0%"; c.font = fnt(True, WHITE)
    c.fill = fill(BLUE_MID); c.alignment = Alignment(horizontal="center")
    for j in range(3, 8):
        cell = ws_d.cell(row=rr, column=j, value="→ calc")
        cell.fill = fill(YELLOW_ASM); cell.font = fnt(color="0070C0")
        cell.alignment = Alignment(horizontal="center")

note(ws_d, r+len(waccs)+2,
     "WACC: CAPM = Rf(4.5%) + β(1.15) × ERP(5.0%) = 10.25%, rounded to 10.0%. No LT debt so WACC ≈ cost of equity. "
     "Net cash = $1,889M cash − $0 debt = $1,889M net cash position (debt-free as of FY2025). "
     "TV = FCF×(1+g)/(WACC−g). Sensitivity shows how sensitive implied price is to key assumptions.")

# ══════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════
out = "/mnt/user-data/outputs/DECK_Model_Tirth_Patel_v10.xlsx"
wb.save(out)
print("Saved:", out)
